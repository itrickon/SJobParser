"""
Скрипт для парсинга вакансий с сайта.
Читает URL из XLSX файла, переходит по ссылкам, извлекает данные о вакансиях и контактах.
Использует Playwright для автоматизации браузера.
"""

import time
import re
import sys
import threading
from pathlib import Path
from playwright.sync_api import sync_playwright, TimeoutError as PlaywrightTimeoutError
from openpyxl import load_workbook, Workbook
from openpyxl.styles import Font, Alignment


class VacancyParser:
    def __init__(self, input_file, output_file=None, gui_mode=False, stop_callback=None):
        """
        Инициализация парсера
        """
        self.input_file = Path(input_file).resolve()

        # Проверяем существование файла
        if not self.input_file.exists():
            raise FileNotFoundError(
                f"Файл не найден: {self.input_file}\n"
                f"Текущая директория: {Path.cwd()}\n"
                f"Попробуйте указать полный путь к файлу."
            )

        # Проверяем расширение файла
        if self.input_file.suffix.lower() != '.xlsx':
            raise ValueError(f"Файл должен иметь расширение .xlsx, получен: {self.input_file.suffix}")

        if output_file:
            self.output_file = Path(output_file).resolve()
        else:
            # Если выходной файл не указан, создаем его рядом с входным
            self.output_file = self.input_file.parent / f"{self.input_file.stem}_output.xlsx"

        self.browser = None
        self.page = None
        self.gui_mode = gui_mode
        self.gui_works = gui_mode  # Для совместимости с trigger_enter_from_gui
        self.stop_callback = stop_callback  # Callback для проверки остановки
        # Событие для ожидания Enter из GUI
        self.enter_event = threading.Event() if gui_mode else None

    def is_stopped(self):
        """Проверка флага остановки (для GUI)"""
        if self.stop_callback:
            try:
                return self.stop_callback()
            except Exception:
                return False
        return False
    
    def read_urls_from_excel(self):
        """Чтение URL из первого столбца Excel файла"""
        try:
            print(f"Открытие файла: {self.input_file}")
            wb = load_workbook(self.input_file)
            ws = wb.active
            urls = []
            
            print(f"Чтение данных из листа '{ws.title}'")
            for row_num, row in enumerate(ws.iter_rows(min_row=1, max_col=1, values_only=True), start=1):
                url = row[0]
                if url:
                    url_str = str(url).strip()
                    # Пропускаем заголовки
                    if url_str.lower() in ['url', 'ссылка', 'link', 'адрес']:
                        print(f"  Пропущен заголовок в строке {row_num}: '{url_str}'")
                        continue
                    # Проверяем, что это URL
                    if isinstance(url, str) and (url_str.startswith('http://') or url_str.startswith('https://')):
                        urls.append(url_str)
                    elif url_str:
                        print(f"  Пропущена строка {row_num}: '{url_str}' (не является URL)")
            
            print(f"Найдено {len(urls)} валидных URL")
            return urls
        except FileNotFoundError as e:
            print(f"ОШИБКА: Файл не найден: {e}")
            print(f"Проверьте путь к файлу: {self.input_file}")
            return []
        except Exception as e:
            print(f"ОШИБКА при чтении файла: {e}")
            print(f"Убедитесь, что файл '{self.input_file}' является валидным Excel файлом (.xlsx)")
            return []
    
    def find_phone_block(self):
        """Поиск блока с номером телефона на +7 (включая замаскированные с ••••)"""
        try:
            # Находим позицию блока "Похожие вакансии"
            similar_block = self.page.locator("xpath=//*[contains(text(), 'Похожие вакансии')]").first
            similar_y = None
            
            if similar_block.count() > 0:
                try:
                    similar_box = similar_block.bounding_box()
                    if similar_box:
                        similar_y = similar_box['y']
                except:
                    pass
            
            # Сначала ищем замаскированные телефоны (формат: +7 XXX XXX ••••)
            masked_phone_pattern = r'\+7\s*\d{1,3}\s*\d{1,3}\s*•{2,}'

            # Ищем элементы с замаскированным телефоном (SuperJob формат)
            masked_selectors = [
                # Кнопки и ссылки с телефоном (основные селекторы SuperJob)
                "xpath=//button[contains(text(), '+7') and contains(text(), '•')]",
                "xpath=//a[contains(text(), '+7') and contains(text(), '•')]",
                "xpath=//span[contains(text(), '+7') and contains(text(), '•')]",
                # С классами SuperJob
                "xpath=//a[contains(@class, 'f-test-link-phone') and contains(text(), '+7')]",
                "xpath=//a[contains(@class, 'ST2tq') and contains(text(), '+7') and contains(text(), '•')]",
                "xpath=//button[contains(@class, 'f-test-button-show-phone') and contains(text(), '+7')]",
            ]

            for selector in masked_selectors:
                try:
                    elements = self.page.locator(selector).all()
                    for elem in elements:
                        if elem.is_visible():
                            # Проверяем позицию элемента - должен быть ВЫШЕ "Похожие вакансии"
                            if similar_y is not None:
                                try:
                                    elem_box = elem.bounding_box()
                                    if elem_box and elem_box['y'] >= similar_y:
                                        continue  # Пропускаем элементы в блоке "Похожие вакансии"
                                except:
                                    pass
                            
                            text = elem.text_content() or ''
                            if re.search(masked_phone_pattern, text):
                                return elem
                            else:
                                print(f"  [DEBUG] Элемент найден, но текст не подходит: {text[:30]}")
                except Exception as e:
                    print(f"  [DEBUG] Ошибка селектора {selector[:50]}...: {e}")
                    continue

            # Ищем полные телефоны (если уже раскрыты)
            phone_pattern = r'\+7\s*\d{1,3}\s*\d{1,3}\s*\d{1,4}\s*\d{1,4}'

            # Ищем ссылки с tel:
            tel_links = self.page.locator("a[href^='tel:']").all()
            for link in tel_links:
                if link.is_visible():
                    # Проверяем позицию
                    if similar_y is not None:
                        try:
                            link_box = link.bounding_box()
                            if link_box and link_box['y'] >= similar_y:
                                continue  # Пропускаем элементы в блоке "Похожие вакансии"
                        except:
                            pass
                    
                    href = link.get_attribute('href') or ''
                    if re.search(phone_pattern, href):
                        return link

            # Ищем элементы с текстом +7 (полный номер)
            elements = self.page.locator("xpath=//*[contains(text(), '+7')]").all()
            for element in elements:
                if element.is_visible():
                    # Проверяем позицию
                    if similar_y is not None:
                        try:
                            elem_box = element.bounding_box()
                            if elem_box and elem_box['y'] >= similar_y:
                                continue  # Пропускаем элементы в блоке "Похожие вакансии"
                        except:
                            pass
                    
                    text = element.text_content() or ''
                    if re.search(phone_pattern, text) and '•' not in text:
                        return element

        except Exception as e:
            print(f"Ошибка при поиске блока с телефоном: {e}")

        return None
    
    
    def extract_vacancy_info(self):
        """Извлечение информации о вакансии с основной страницы"""
        vacancy_title = ""
        company_name = ""
        address = ""
        experience = ""
        
        try:
            # Название вакансии (обычно в заголовке)
            title_selectors = [
                "xpath=//h1[contains(@data-qa, 'vacancy-title')]",
                "xpath=//h1[contains(@class, 'vacancy-title')]",
                "xpath=//h1",  
                "xpath=//h2[contains(@class, 'title')]",
                "xpath=//*[contains(@class, 'vacancy-title')]",
                "xpath=//*[contains(@class, 'title')]",
                "xpath=//div[contains(@class, 'header')]//h1",
                "xpath=//div[contains(@class, 'header')]//h2"
            ]
            
            for selector in title_selectors:
                try:
                    title_elem = self.page.locator(selector).first
                    if title_elem.is_visible():
                        title_text = (title_elem.text_content() or '').strip()
                        # Убираем зарплату если она в заголовке
                        title_text = re.sub(r'\s*до\s*\d+\s*₽.*$', '', title_text)
                        title_text = re.sub(r'\s*от\s*\d+\s*₽.*$', '', title_text)
                        if title_text and len(title_text) > 3:
                            vacancy_title = title_text
                            break
                except:
                    continue
            
            # Название компании (обновленные селекторы для SuperJob)
            # Ищем элемент с классом _2alGT (border: 1px solid) и берем первый текстовый элемент
            company_selectors = [
                # SuperJob - элемент с border (класс _2alGT)
                "xpath=//div[contains(@class, '_2alGT')]//a[3]",
                "xpath=//div[contains(@class, '_2alGT')]//span[3]",
                "xpath=//div[contains(@class, '_2alGT')]//*[not(*)][normalize-space(text())][3]",
                # Альтернативные селекторы для блока _3fYQD
                "xpath=//div[contains(@class, '_3fYQD')]//a[contains(@href, '/clients/')][1]",
                "xpath=//div[contains(@class, '_3fYQD')]//a[contains(@href, 'vacancies')][1]",
                "xpath=//div[contains(@class, '_3fYQD')]//div[contains(@class, '_3YL-9')]//a[1]",
                "xpath=//div[contains(@class, '_3fYQD')]//div[contains(@class, '_3S0Ir')]//a[1]",
                # Ссылки с /clients/ (страница компании)
                "xpath=//a[contains(@href, '/clients/') and contains(@class, 'f-test-link')][1]",
                "xpath=//a[contains(@href, '/clients/') and not(contains(@href, 'reviews'))][1]",
                # Старые селекторы (для совместимости)
                "xpath=//a[contains(@data-qa, 'vacancy-company-name')]",
                "xpath=//*[contains(@data-qa, 'vacancy-company-name')]",
                "xpath=//h2[contains(@class, 'company')]//a",
                "xpath=//h2[contains(@class, 'company')]",
                "xpath=//*[contains(@class, 'company')]//a",
                "xpath=//*[contains(@class, 'employer')]//a",
                "xpath=//*[contains(@class, 'company')]",
                "xpath=//*[contains(@class, 'employer')]",
                "xpath=//a[contains(@class, 'company')]",
                "xpath=//span[contains(@class, 'company')]"
            ]

            for selector in company_selectors:
                try:
                    company_elems = self.page.locator(selector).all()
                    for company_elem in company_elems:
                        if company_elem.is_visible():
                            company_text = (company_elem.text_content() or '').strip()
                            if company_text and len(company_text) > 1:
                                # Проверяем, что это не ссылка или служебный текст
                                if not company_text.startswith('http') and len(company_text) < 100:
                                    # Убираем служебные фразы
                                    company_text = re.sub(r'\s*Проверенный работодатель.*$', '', company_text)
                                    company_text = re.sub(r'\s*Клиент.*$', '', company_text)
                                    company_text = re.sub(r'\s*Рекрутер.*$', '', company_text)
                                    # Проверяем, что это не название вакансии
                                    if company_text.lower() != vacancy_title.lower():
                                        company_name = company_text
                                        break
                    if company_name:
                        break
                except:
                    continue
            
            # Адрес (SuperJob формат)
            address_selectors = [
                "xpath=//span[contains(@data-qa, 'vacancy-view-raw-address')]",
                "xpath=//*[contains(@data-qa, 'vacancy-view-location')]",
                "xpath=//*[contains(@class, 'address')]",
                "xpath=//*[contains(@class, 'location')]",
                "xpath=//*[contains(@class, 'geo')]",
                "xpath=//*[contains(text(), 'адрес')]/following-sibling::*[1]",
                "xpath=//*[contains(text(), 'Адрес')]/following-sibling::*[1]",
                "xpath=//*[contains(text(), 'адрес')]/parent::*",
                "xpath=//*[contains(text(), 'Адрес')]/parent::*"
            ]
            
            for selector in address_selectors:
                try:
                    addr_elems = self.page.locator(selector).all()
                    for addr_elem in addr_elems:
                        if addr_elem.is_visible():
                            addr_text = (addr_elem.text_content() or '').strip()
                            # Убираем слово "Адрес:" если есть
                            addr_text = re.sub(r'^[Аа]дрес\s*:?\s*', '', addr_text)
                            addr_text = re.sub(r'\s*Показать на карте.*$', '', addr_text)
                            if addr_text and len(addr_text) > 2 and not addr_text.startswith('http'):
                                # Проверяем, что это похоже на адрес (содержит буквы или цифры)
                                if re.search(r'[А-Яа-яA-Za-z0-9]', addr_text):
                                    address = addr_text
                                    break
                    if address:
                        break
                except:
                    continue
            
            # Опыт работы (SuperJob формат: "Опыт работы от 3 лет, высшее образование")
            experience_selectors = [
                "xpath=//*[contains(@data-qa, 'vacancy-experience')]",
                "xpath=//*[contains(@class, 'experience')]",
                "xpath=//*[contains(text(), 'Опыт работы')]",  # SuperJob формат
                "xpath=//*[contains(text(), 'опыт работы')]",
                "xpath=//*[contains(text(), 'опыт')]/following-sibling::*[1]",
                "xpath=//*[contains(text(), 'Опыт')]/following-sibling::*[1]",
                "xpath=//*[contains(text(), 'опыт')]/parent::*",
                "xpath=//*[contains(text(), 'Опыт')]/parent::*"
            ]
            
            for selector in experience_selectors:
                try:
                    exp_elems = self.page.locator(selector).all()
                    for exp_elem in exp_elems:
                        if exp_elem.is_visible():
                            exp_text = (exp_elem.text_content() or '').strip()
                            # Проверяем наличие ключевых слов
                            if any(keyword in exp_text.lower() for keyword in ['опыт', 'experience', 'стаж', 'лет', 'год']):
                                # Убираем слово "Опыт работы:" если есть
                                exp_text = re.sub(r'^[Оо]пыт\s+работы\s*:?\s*', '', exp_text)
                                exp_text = re.sub(r'^[Оо]пыт\s*:?\s*', '', exp_text)
                                # Берем только часть про опыт (до запятой или точки с запятой)
                                exp_text = re.split(r'[,;]', exp_text)[0].strip()
                                if 'div' in exp_text:
                                    exp_text = ''
                                if exp_text and len(exp_text) > 1:
                                    experience = exp_text
                                    break
                    if experience:
                        break
                except:
                    continue
                    
        except Exception as e:
            print(f"Ошибка при извлечении информации о вакансии: {e}")
            
        return vacancy_title, company_name, address, experience
    
    def parse_vacancy(self, url):
        """Парсинг одной вакансии"""
        print(f"Обработка: {url}")
        
        try:
            # Пробуем загрузить страницу с несколькими стратегиями ожидания
            try:
                self.page.goto(url, wait_until='domcontentloaded', timeout=30000)
            except PlaywrightTimeoutError:
                # Если не удалось загрузить, пробуем с другой стратегией
                try:
                    self.page.goto(url, wait_until='load', timeout=20000)
                except PlaywrightTimeoutError:
                    print(f"  Предупреждение: страница загружается медленно, продолжаем...")
                    self.page.goto(url, timeout=20000)
            time.sleep(0.7)  # Дополнительная пауза для загрузки
            
            # Извлекаем информацию о вакансии с основной страницы
            vacancy_title, company_name, address, experience = self.extract_vacancy_info()
            
            print(f"  Найдено: Вакансия='{vacancy_title}', Компания='{company_name}'")
            
            # Ищем блок с телефоном и кликаем на него
            phone_block = self.find_phone_block()
            name = ""
            phone = ""
            
            if phone_block:
                try:
                    phone_text = phone_block.text_content() or phone_block.get_attribute('href') or ''
                    print(f"  Найден блок с телефоном: {phone_text[:30]}, кликаем...")

                    # Прокручиваем к элементу
                    phone_block.scroll_into_view_if_needed()
                    time.sleep(0.6)

                    # Кликаем на блок с телефоном
                    phone_block.click(timeout=5000)
                    
                    # Ждем появления модального окна с контактами (до 10 секунд)
                    modal_found = False
                    for attempt in range(10):
                        try:
                            # Ищем модальное окно по тексту "Вы обменялись контактами"
                            modal = self.page.locator("xpath=//*[contains(text(), 'Вы обменялись контактами')]/ancestor::div[contains(@class, 'ltQhb')]").first
                            if modal.count() > 0:
                                modal.wait_for(state='visible', timeout=500)
                                modal_found = True
                                break
                        except Exception:
                            pass
                        time.sleep(0.5)
                    
                    if modal_found:
                        try:
                            # Ждем появления телефона (ссылка с href="tel:")
                            phone_elem = modal.locator("xpath=//a[contains(@href, 'tel:')]").first
                            
                            # Ждем до 5 секунд появления телефона
                            for attempt in range(10):
                                try:
                                    if phone_elem.count() > 0 and phone_elem.is_visible():
                                        phone = (phone_elem.text_content() or '').strip()
                                        if phone and '+' in phone:
                                            break
                                except Exception:
                                    pass
                                time.sleep(0.5)
                            
                            # Ищем имя (класс wyL3A или _2Yqdk)
                            if not phone:
                                name_elem = modal.locator("xpath=//span[contains(@class, 'wyL3A')]").first
                                if name_elem.count() > 0 and name_elem.is_visible():
                                    name = (name_elem.text_content() or '').strip()
                            
                            if phone:
                                print(f"  Контакты: Имя='{name}', Телефон='{phone}'")
                            elif name and name not in ['ошибка', 'контактов', 'количество']:
                                print(f"  Контакты: Имя='{name}', Телефон='{phone}'")
                            else:
                                print(f'  Проблемы с контактами (модальное окно найдено, но телефон не появился)')
                        except Exception as e:
                            print(f"  Ошибка при извлечении контактов: {e}")
                    else:
                        print("  Модальное окно с контактами не найдено (таймаут ожидания)")
                    
                    # Закрываем модальное окно (если есть кнопка закрытия)
                    try:
                        close_selectors = [
                            "xpath=//button[contains(@class, 'close')]",
                            "xpath=//button[contains(@class, '_3inGK')]",
                            "xpath=//button[@title='']",
                            "xpath=//button[@aria-label='Закрыть']",
                            "xpath=//*[contains(@class, 'close')]//button",
                            "xpath=//*[contains(@class, 'close')]"
                        ]
                        for selector in close_selectors:
                            try:
                                close_btn = self.page.locator(selector).first
                                if close_btn.is_visible(timeout=1000):
                                    close_btn.click()
                                    time.sleep(0.5)
                                    break
                            except:
                                continue
                    except:
                        pass
                except Exception as e:
                    print(f"  Ошибка при клике на блок с телефоном: {e}")
            else:
                print("  Блок с телефоном не найден")
                
            error_inb = lambda x: x not in name.lower()
            # Проверяем, не превышен ли лимит контактов
            limit_element = self.page.get_by_text("Максимальное количество контактов")
            if phone and limit_element.count() == 0 and error_inb('ошибка') and error_inb('контактов') and error_inb('количество'):
                return {
                    'url': url,
                    'vacancy_title': vacancy_title,
                    'company_name': company_name,
                    'name': name,
                    'phone': phone,
                    'address': address,
                    'experience': experience
                }
            elif limit_element.count() > 0:
                print("  Превышен лимит контактов")
                return None
            else:
                return None
            
        except Exception as e:
            print(f"  ОШИБКА при парсинге: {e}")
            return {
                'url': url,
                'vacancy_title': '',
                'company_name': '',
                'name': '',
                'phone': '',
                'address': '',
                'experience': ''
            }
    
    def trigger_enter_from_gui(self):
        """Вызывается из GUI для имитации нажатия Enter"""
        if self.gui_works and self.enter_event:
            print("GUI: Вход подтвержден, устанавливаем событие Enter")
            self.enter_event.set()
            print(f"GUI: Событие установлено, is_set={self.enter_event.is_set()}")
    
    def save_to_excel(self, data):
        """Сохранение данных в Excel файл"""
        wb = Workbook()
        ws = wb.active
        ws.title = "Вакансии"
        
        # Заголовки
        headers = ['URL', 'Название вакансии', 'Название компании', 'Имя', 'Телефон', 'Адрес', 'Опыт']
        ws.append(headers)
        
        # Стили для заголовков
        header_font = Font(bold=True, size=12)
        header_alignment = Alignment(horizontal='center', vertical='center')
        
        for cell in ws[1]:
            cell.font = header_font
            cell.alignment = header_alignment
        
        # Данные
        for row_data in data:
            ws.append([
                row_data.get('url', ''),
                row_data.get('vacancy_title', ''),
                row_data.get('company_name', ''),
                row_data.get('name', ''),
                row_data.get('phone', ''),
                row_data.get('address', ''),
                row_data.get('experience', '')
            ])
        
        # Автоподбор ширины столбцов
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width
        
        wb.save(self.output_file)
        print(f"\nДанные сохранены в файл: {self.output_file}")
    
    def run(self):
        """Запуск парсера"""
        print("Инициализация браузера...")
        
        with sync_playwright() as p:
            # Запускаем браузер
            self.browser = p.chromium.launch(
                headless=False,  # Показываем браузер для отладки
                args=['--start-maximized']
            )
            
            # Создаем контекст с настройками
            context = self.browser.new_context(
                viewport={'width': 900, 'height': 700},
                user_agent='Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36'
            )
            
            self.page = context.new_page()

            try:
                # Открываем страницу SuperJob для входа в аккаунт
                print("\n" + "="*60)
                print("Открываю страницу SuperJob для входа в аккаунт...")
                print("="*60)
                self.page.goto("https://ramenskoe.superjob.ru/?curtain%5BreturnUrl%5D=%2F&curtain%5BrouteName%5D=authLogin&curtain%5Bid%5D=auth", wait_until='domcontentloaded', timeout=30000)

                print("\nВАЖНО: Войдите в свой аккаунт SuperJob в открывшемся браузере")
                
                # Ждем нажатия Enter от пользователя
                if self.gui_mode:
                    print("   После входа в аккаунт нажмите кнопку 'Вход выполнен' в GUI")
                    print("   (У вас есть неограниченное время для входа)\n")
                    self.enter_event.wait()  # Ждем события от GUI
                    self.enter_event.clear()  # Сбрасываем для следующего использования
                    print("\n Вход подтвержден через GUI. Начинаю парсинг...\n")
                else:
                    print("   После входа в аккаунт вернитесь сюда и нажмите ENTER для продолжения...")
                    print("   (У вас есть неограниченное время для входа)\n")
                    input("Нажмите ENTER после входа в аккаунт для начала парсинга... ")
                    print("\n Вход подтвержден. Начинаю парсинг...\n")
                
                # Читаем URL из файла
                print("Чтение URL из файла...")
                urls = self.read_urls_from_excel()
                
                if not urls:
                    print("Не найдено URL в файле!")
                    return
                
                print(f"Найдено {len(urls)} URL для обработки\n")

                results = []
                for i, url in enumerate(urls, 1):
                    # Проверяем флаг остановки
                    if self.is_stopped():
                        print("\n\nПарсинг остановлен пользователем")
                        break

                    print(f"[{i}/{len(urls)}] ", end="")
                    result = self.parse_vacancy(url)
                    if result:
                        results.append(result)
                    time.sleep(1.2)  # Пауза между запросами

                print("\nСохранение результатов...")
                if results:
                    self.save_to_excel(results)
                else:
                    print("Нет результатов для сохранения")

            finally:
                print("\nЗакрытие браузера...")
                if self.gui_mode:
                    # В режиме GUI закрываем браузер без ожидания
                    # Браузер может быть уже закрыт через taskkill
                    try:
                        if self.browser and self.browser.is_connected():
                            self.browser.close()
                            print("Браузер закрыт успешно")
                        else:
                            print("Браузер уже закрыт")
                    except Exception as e:
                        # Игнорируем ошибки если браузер уже закрыт
                        print(f"Браузер закрыт (или уже был закрыт): {type(e).__name__}")
                else:
                    input("Нажмите ENTER для закрытия браузера... ")
                    self.browser.close()


def main():
    try:
        parser = VacancyParser(input_file="superjob_vacancies.xlsx", output_file="output.xlsx")
        parser.run()
    except FileNotFoundError as e:
        print(f"\n{e}")
        print("\nПодсказки:")
        print("  - Убедитесь, что файл существует")
        print("  - Проверьте правильность пути к файлу")
        print("  - Можно использовать относительный путь (от текущей директории)")
        print("  - Или абсолютный путь (полный путь к файлу)")
        sys.exit(1)
    except KeyboardInterrupt:
        print("\n\nПарсинг прерван пользователем")
        sys.exit(0)
    except Exception as e:
        print(f"\nПроизошла ошибка: {e}")
        sys.exit(1)


if __name__ == "__main__":
    main()
