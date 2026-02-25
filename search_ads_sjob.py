"""
Парсер вакансий с SuperJob.ru
Образовательное использование. Соблюдайте правила сайта и robots.txt.
"""

import os
import re
import asyncio
from urllib.parse import urlparse, urlencode, parse_qsl, unquote
from playwright.async_api import async_playwright
from openpyxl import Workbook, load_workbook


class SearchSuperJob:
    def __init__(self, url: str, max_vacancies: int = 100):
        """
        :param url: URL страницы поиска
        :param max_vacancies: Максимальное количество вакансий для сбора
        """
        self.base_url = self._normalize_url(url.rstrip("/"))
        # Убираем page= из URL для формирования пагинации
        if "&page=" in self.base_url:
            self.base_url = self.base_url.split("&page=")[0]
        elif "?page=" in self.base_url:
            parts = self.base_url.split("?page=")
            self.base_url = parts[0]
        self.max_vacancies = max_vacancies
        self.vacancies = []
        self.data_saving = "superjob_parse_results/superjob_vacancies.xlsx"
        self.warning_message()

    def _normalize_url(self, url: str) -> str:
        """Исправление двойной/некорректной кодировки в URL."""
        try:
            parsed = urlparse(url)
            if not parsed.query:
                return url
            params = parse_qsl(parsed.query, keep_blank_values=True)
            new_params = []
            for k, v in params:
                if v and k == "keywords":
                    v = self._decode_keywords(v)  # raw string, urlencode закодирует
                new_params.append((k, v))
            new_query = urlencode(new_params, doseq=True, encoding="utf-8")
            return f"{parsed.scheme}://{parsed.netloc}{parsed.path}?{new_query}"
        except Exception:
            return url

    def _decode_keywords(self, value: str) -> str:
        """Декодирование keywords с исправлением двойной кодировки и mojibake."""
        prev, decoded = None, value
        while prev != decoded:
            prev, decoded = decoded, unquote(decoded, encoding="utf-8", errors="replace")
        if self._looks_valid_cyrillic(decoded):
            return decoded
        try:
            return decoded.encode("latin-1").decode("utf-8", errors="replace")
        except (UnicodeEncodeError, UnicodeDecodeError):
            return decoded

    def _looks_valid_cyrillic(self, s: str) -> bool:
        """Проверка на нормальный кириллический текст"""
        if not s or len(s) < 2:
            return False
        cyrillic = sum(1 for c in s if "\u0400" <= c <= "\u04FF")
        non_cyrillic = len(s) - cyrillic - s.count(" ")
        return cyrillic >= 3 and non_cyrillic < len(s) // 2

    async def _get_vacancies_from_page(self):
        """Сбор вакансий с текущей страницы"""
        vacancies = []

        # Ссылки на вакансии: /vakansii/[slug]-[id].html
        # Исключаем фильтры. Вакансии имеют вид: .../vakansii/slug-51689847.html
        link_selector = 'a[href*="/vakansii/"][href*=".html"]'
        links = await self.page.query_selector_all(link_selector)
        filter_patterns = (
            r"vahta\.html", r"\d+x\d+-rabota\.html", r"40tys/", r"50tys/",
            r"nepolnyj-den", r"smennyj-grafik", r"rabota-dlya-pensionerov",
            r"bez-opyta", r"rabota-s-prozhivaniem", r"udalenka",
        )
        is_filter = re.compile("|".join(filter_patterns))
        # Вакансии имеют ID: slug-51689847.html
        has_vacancy_id = re.compile(r"-\d{6,}\.html")
        # Фильтр по ключевым словам (исключаем вакансии с этими словами в URL)
        exclude_words = re.compile(r"voennosluzhaschij|kontrakt", re.IGNORECASE)

        seen_urls = set()
        for link in links:
            try:
                href = await link.get_attribute("href")
                if not href:
                    continue
                if is_filter.search(href) or not has_vacancy_id.search(href):
                    continue
                # Исключаем URL с помощью филотрации
                if exclude_words.search(href):
                    continue
                # Нормализуем URL (могут быть относительные)
                if href.startswith("/"):
                    full_url = f"https://www.superjob.ru{href}"
                elif href.startswith("http"):
                    full_url = href
                else:
                    full_url = f"https://www.superjob.ru/{href}"

                if full_url in seen_urls:
                    continue
                seen_urls.add(full_url)

                vacancies.append({"url": full_url})
            except Exception:
                continue

        return vacancies

    async def _go_to_next_page(self, page_num: int) -> bool:
        """Переход на следующую страницу по URL"""
        sep = "&" if "?" in self.base_url else "?"
        next_url = f"{self.base_url}{sep}page={page_num}"
        try:
            next = await self.page.goto(next_url, wait_until="domcontentloaded", timeout=15000)
            if self._normalize_url(next.url.rstrip("/")) == self.base_url:
                return False
            await asyncio.sleep(1.5)
            return True
            
        except Exception as e:
            return False

    def _create_xlsx(self):
        """Создание XLSX файла с заголовками"""
        os.makedirs(os.path.dirname(self.data_saving), exist_ok=True)
        wb = Workbook()
        ws = wb.active
        ws.title = "SuperJob Вакансии"
        headers = ["Ссылка"]
        for col, header in enumerate(headers, start=1):
            ws.cell(row=1, column=col, value=header)
        wb.save(self.data_saving)
        print(f"Создан файл: {self.data_saving}")

    def _save_to_xlsx(self):
        """Сохранение данных в XLSX"""
        if not os.path.exists(self.data_saving):
            self._create_xlsx()

        wb = load_workbook(self.data_saving)
        ws = wb.active
        start_row = ws.max_row + 1 if ws.max_row > 1 else 2

        for i, v in enumerate(self.vacancies, start=start_row):
            ws.cell(row=i, column=1, value=v.get("url", ""))
            ws.cell(row=i, column=2, value=v.get("title", ""))
            ws.cell(row=i, column=3, value=v.get("salary", ""))
            ws.cell(row=i, column=4, value=v.get("company", ""))
            ws.cell(row=i, column=5, value=v.get("location", ""))

        wb.save(self.data_saving)
        print(f"Данные сохранены: {self.data_saving}")

    def warning_message(self):
        print("\n" + "=" * 50)
        print("EDUCATIONAL USE ONLY - NO WARRANTY PROVIDED")
        print("This parser may violate Terms of Service.")
        print("Use only for learning web scraping techniques.")
        print("Author not responsible for any legal consequences.")
        print("=" * 50 + "\n")

    async def parse_main(self, update_callback=None):
        def log(msg):
            if update_callback:
                update_callback(msg)
            print(msg)

        self._create_xlsx()

        async with async_playwright() as playwright:
            browser = await playwright.chromium.launch(headless=False)
            context = await browser.new_context()
            self.page = await context.new_page()

            try:
                log("Переход на страницу поиска...")
                await self.page.goto(
                    self.base_url,
                    wait_until="domcontentloaded",
                    timeout=60000,
                )
                await asyncio.sleep(2)

                page_num = 1
                empty_pages = 0

                while len(self.vacancies) < self.max_vacancies:
                    log(f"Парсинг страницы {page_num}...")
                    page_vacancies = await self._get_vacancies_from_page()

                    if not page_vacancies:
                        empty_pages += 1
                        if empty_pages >= 2:
                            log("Две пустые страницы подряд. Завершение.")
                            break
                    else:
                        empty_pages = 0
                        for v in page_vacancies:
                            if len(self.vacancies) < self.max_vacancies:
                                self.vacancies.append(v)

                    log(f"Собрано вакансий: {len(self.vacancies)} из {self.max_vacancies}")

                    if len(self.vacancies) >= self.max_vacancies:
                        break

                    page_num += 1
                    if not await self._go_to_next_page(page_num):
                        break

                    await asyncio.sleep(1.5)

                self._save_to_xlsx()
                log(f"Готово. Всего вакансий: {len(self.vacancies)}")

            finally:
                await browser.close()


async def main():
    url = (
        "https://www.superjob.ru/vacancy/search/?keywords=Технический%20писатель%20саратов&geo%5Bt%5D%5B0%5D=4"
    )
    parser = SearchSuperJob(url=url, max_vacancies=50)
    await parser.parse_main()


if __name__ == "__main__":
    asyncio.run(main())
